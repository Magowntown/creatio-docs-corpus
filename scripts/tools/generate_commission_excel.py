#!/usr/bin/env python3
"""Generate Commission Excel report locally by querying Creatio and creating xlsx.

This bypasses the broken IntExcelExport library by:
1. Querying data via Creatio DataService API
2. Generating Excel locally using openpyxl
3. Optionally uploading back to Creatio or saving locally

Usage:
    python3 generate_commission_excel.py --year-month 2025-01 --output report.xlsx
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import requests

# Add repo root to path
REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

from scripts._env import load_dotenv
from scripts._paths import ARTIFACTS_DIR, ensure_dirs

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

load_dotenv()
ensure_dirs()

CREATIO_URL = os.environ.get("CREATIO_URL", "https://dev-pampabay.creatio.com").rstrip("/")
USERNAME = os.environ.get("CREATIO_USERNAME", "")
PASSWORD = os.environ.get("CREATIO_PASSWORD", "")
TIMEOUT = int(os.environ.get("CREATIO_TIMEOUT_SECONDS", "120"))


def login(session):
    """Login to Creatio and return headers with CSRF token."""
    resp = session.post(
        f"{CREATIO_URL}/ServiceModel/AuthService.svc/Login",
        json={"UserName": USERNAME, "UserPassword": PASSWORD},
        timeout=TIMEOUT,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Login failed: {resp.status_code}")

    return {
        "Content-Type": "application/json",
        "BPMCSRF": session.cookies.get("BPMCSRF", ""),
    }


def get_year_month_id(session, headers, year_month_name):
    """Get BGYearMonth ID by name (e.g., '2025-01')."""
    esq = {
        "rootSchemaName": "BGYearMonth",
        "operationType": 0,
        "columns": {"items": {"Id": {"expression": {"columnPath": "Id"}}}},
        "filters": {
            "filterType": 6,
            "items": {
                "Filter": {
                    "filterType": 1,
                    "comparisonType": 3,
                    "leftExpression": {"expressionType": 0, "columnPath": "Name"},
                    "rightExpression": {
                        "expressionType": 2,
                        "parameter": {"dataValueType": 1, "value": year_month_name},
                    },
                }
            },
        },
    }

    url = f"{CREATIO_URL}/0/DataService/json/SyncReply/SelectQuery"
    resp = session.post(url, json=esq, headers=headers, timeout=TIMEOUT)
    data = resp.json()
    rows = data.get("rows", [])
    return rows[0]["Id"] if rows else None


def query_commission_data(session, headers, year_month_name=None, sales_group_id=None):
    """Query BGCommissionReportDataView with date-based filtering."""

    # Build ESQ - skip BGYearMonth columns to avoid INNER JOIN issues
    esq = {
        "rootSchemaName": "BGCommissionReportDataView",
        "operationType": 0,
        "columns": {
            "items": {
                "Id": {"expression": {"columnPath": "Id"}},
                "BGSalesRep": {"expression": {"columnPath": "BGSalesRep"}},
                "BGSalesItem": {"expression": {"columnPath": "BGSalesItem"}},
                "BGTransactionDate": {"expression": {"columnPath": "BGTransactionDate"}},
                "BGCommissionAmount": {"expression": {"columnPath": "BGCommissionAmount"}},
                "BGSalesAmount": {"expression": {"columnPath": "BGSalesAmount"}},
                "BGSalesRepName": {"expression": {"columnPath": "BGSalesRep.Name"}},
            }
        },
        "filters": {"filterType": 6, "items": {}},
    }

    filter_idx = 0

    # Date-based filter (FLT-004 fix)
    if year_month_name:
        try:
            year, month = map(int, year_month_name.split("-"))
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)

            esq["filters"]["items"][f"DateGte_{filter_idx}"] = {
                "filterType": 1,
                "comparisonType": 4,  # >=
                "leftExpression": {"expressionType": 0, "columnPath": "BGTransactionDate"},
                "rightExpression": {
                    "expressionType": 2,
                    "parameter": {
                        "dataValueType": 7,
                        "value": start_date.strftime("%Y-%m-%dT00:00:00.000Z"),
                    },
                },
            }
            filter_idx += 1

            esq["filters"]["items"][f"DateLt_{filter_idx}"] = {
                "filterType": 1,
                "comparisonType": 6,  # <
                "leftExpression": {"expressionType": 0, "columnPath": "BGTransactionDate"},
                "rightExpression": {
                    "expressionType": 2,
                    "parameter": {
                        "dataValueType": 7,
                        "value": end_date.strftime("%Y-%m-%dT00:00:00.000Z"),
                    },
                },
            }
            filter_idx += 1
        except ValueError:
            print(f"WARN: Could not parse year-month: {year_month_name}")

    # Sales group filter
    if sales_group_id:
        esq["filters"]["items"][f"SalesGroup_{filter_idx}"] = {
            "filterType": 1,
            "comparisonType": 3,
            "leftExpression": {"expressionType": 0, "columnPath": "BGSalesRep.BGSalesGroupLookup"},
            "rightExpression": {
                "expressionType": 2,
                "parameter": {"dataValueType": 0, "value": sales_group_id},
            },
        }

    url = f"{CREATIO_URL}/0/DataService/json/SyncReply/SelectQuery"
    resp = session.post(url, json=esq, headers=headers, timeout=TIMEOUT)
    data = resp.json()
    return data.get("rows", [])


def create_excel(rows, output_path):
    """Create Excel file from data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Headers
    headers = ["Sales Rep", "Sales Item", "Transaction Date", "Commission Amount", "Sales Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("BGSalesRepName") or row.get("BGSalesRep", {}).get("displayValue", ""))
        ws.cell(row=row_idx, column=2, value=row.get("BGSalesItem", ""))

        # Parse date
        date_val = row.get("BGTransactionDate")
        if date_val:
            try:
                dt = datetime.fromisoformat(date_val.replace("Z", "+00:00"))
                ws.cell(row=row_idx, column=3, value=dt.date())
            except:
                ws.cell(row=row_idx, column=3, value=date_val)

        ws.cell(row=row_idx, column=4, value=row.get("BGCommissionAmount"))
        ws.cell(row=row_idx, column=5, value=row.get("BGSalesAmount"))

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_path)
    return len(rows)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Generate Commission Excel report")
    parser.add_argument("--year-month", default=None, help="Filter by year-month (e.g., 2025-01)")
    parser.add_argument("--sales-group", default=None, help="Filter by sales group ID")
    parser.add_argument("--output", default=None, help="Output file path")
    args = parser.parse_args()

    if not USERNAME or not PASSWORD:
        print("ERROR: Set CREATIO_USERNAME and CREATIO_PASSWORD")
        sys.exit(1)

    session = requests.Session()

    print(f"=== Connecting to {CREATIO_URL} ===")
    headers = login(session)
    print("Login: OK")

    print(f"\n=== Querying Commission data ===")
    if args.year_month:
        print(f"Year-Month filter: {args.year_month}")
    if args.sales_group:
        print(f"Sales Group filter: {args.sales_group}")

    rows = query_commission_data(session, headers, args.year_month, args.sales_group)
    print(f"Retrieved: {len(rows)} rows")

    # Generate output filename
    if args.output:
        output_path = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ym_suffix = f"_{args.year_month}" if args.year_month else ""
        output_path = ARTIFACTS_DIR / f"commission_report{ym_suffix}_{timestamp}.xlsx"

    print(f"\n=== Generating Excel ===")
    row_count = create_excel(rows, output_path)
    print(f"Created: {output_path}")
    print(f"Rows: {row_count}")


if __name__ == "__main__":
    main()
