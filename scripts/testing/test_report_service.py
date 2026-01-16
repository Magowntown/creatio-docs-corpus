#!/usr/bin/env python3
"""End-to-end test for the Creatio Excel report flow in DEV.

This script:
- Logs in
- Resolves the report (IntExcelReport) by IntName unless CREATIO_REPORT_ID is provided
- Resolves BGYearMonth by Name unless CREATIO_YEAR_MONTH_NAME=__NONE__
- Calls POST /0/rest/UsrExcelReportService/Generate
- Downloads via GET /0/rest/UsrExcelReportService/GetReport/{key}/{reportNameSegment} (preferred; serves .xlsm)
  - Falls back to IntExcelReportService/GetReport for older deployments
- Saves the .xlsm under test-artifacts/
- Optionally inspects the workbook (openpyxl) to count non-empty rows

Environment variables:
- CREATIO_URL, CREATIO_USERNAME, CREATIO_PASSWORD
- CREATIO_REPORT_CODE (default: Commission)
- CREATIO_REPORT_ID (optional override)
- CREATIO_REPORT_INT_NAME (optional override)
- CREATIO_REPORT_NAME_SEGMENT (optional override)
- CREATIO_YEAR_MONTH_NAME (default: 2023-05; set to __NONE__ to skip)
- CREATIO_SALES_GROUP_ID (optional)
- CREATIO_TIMEOUT_SECONDS (default: 120)

Note: the frontend uses the field name SalesRepId to carry a SalesGroupId (legacy naming).
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts._env import load_dotenv
from scripts._paths import ARTIFACTS_DIR, ensure_dirs

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None

EMPTY_GUID = "00000000-0000-0000-0000-000000000000"

load_dotenv()
ensure_dirs()

CREATIO_URL = os.environ.get("CREATIO_URL", "https://dev-pampabay.creatio.com").rstrip("/")
USERNAME = os.environ.get("CREATIO_USERNAME", "")
PASSWORD = os.environ.get("CREATIO_PASSWORD", "")

REPORT_CODE = os.environ.get("CREATIO_REPORT_CODE", "Commission").strip() or "Commission"
REPORT_ID_OVERRIDE = os.environ.get("CREATIO_REPORT_ID", "").strip()
REPORT_INT_NAME_OVERRIDE = os.environ.get("CREATIO_REPORT_INT_NAME", "").strip()
REPORT_NAME_SEGMENT_OVERRIDE = os.environ.get("CREATIO_REPORT_NAME_SEGMENT", "").strip()

YEAR_MONTH_NAME = os.environ.get("CREATIO_YEAR_MONTH_NAME", "2023-05").strip()
SALES_GROUP_ID = os.environ.get("CREATIO_SALES_GROUP_ID", "").strip()

TIMEOUT_SECONDS = int(os.environ.get("CREATIO_TIMEOUT_SECONDS", "120"))


def _headers(session: requests.Session) -> Dict[str, str]:
    return {
        "Content-Type": "application/json",
        "BPMCSRF": session.cookies.get("BPMCSRF", ""),
    }


def _select_query(session: requests.Session, headers: Dict[str, str], body: Dict[str, Any]) -> Dict[str, Any]:
    url = f"{CREATIO_URL}/0/DataService/json/SyncReply/SelectQuery"
    resp = session.post(url, json=body, headers=headers, timeout=TIMEOUT_SECONDS)
    if resp.status_code != 200:
        raise RuntimeError(f"SelectQuery failed: {resp.status_code} {resp.text[:300]}")
    return resp.json()


def _resolve_report_int_name(report_code: str) -> str:
    if REPORT_INT_NAME_OVERRIDE:
        return REPORT_INT_NAME_OVERRIDE
    # Known DEV naming conventions.
    if report_code == "Commission":
        return "Rpt Commission"
    # IW report uses exact IntName in DEV.
    return report_code


def _resolve_int_excel_report(
    session: requests.Session,
    headers: Dict[str, str],
    report_code: str,
) -> Dict[str, Any]:
    if REPORT_ID_OVERRIDE:
        # Fetch metadata by Id.
        esq_body = {
            "rootSchemaName": "IntExcelReport",
            "operationType": 0,
            "columns": {
                "items": {
                    "Id": {"expression": {"columnPath": "Id"}},
                    "IntName": {"expression": {"columnPath": "IntName"}},
                    "IntEntitySchemaName": {"expression": {"columnPath": "IntEntitySchemaName"}},
                    "IntSheetName": {"expression": {"columnPath": "IntSheetName"}},
                    "IntEsq": {"expression": {"columnPath": "IntEsq"}},
                }
            },
            "filters": {
                "filterType": 6,
                "items": {
                    "IdFilter": {
                        "filterType": 1,
                        "comparisonType": 3,
                        "leftExpression": {"expressionType": 0, "columnPath": "Id"},
                        "rightExpression": {
                            "expressionType": 2,
                            "parameter": {"dataValueType": 0, "value": REPORT_ID_OVERRIDE},
                        },
                    }
                },
            },
        }
    else:
        int_name = _resolve_report_int_name(report_code)
        esq_body = {
            "rootSchemaName": "IntExcelReport",
            "operationType": 0,
            "columns": {
                "items": {
                    "Id": {"expression": {"columnPath": "Id"}},
                    "IntName": {"expression": {"columnPath": "IntName"}},
                    "IntEntitySchemaName": {"expression": {"columnPath": "IntEntitySchemaName"}},
                    "IntSheetName": {"expression": {"columnPath": "IntSheetName"}},
                    "IntEsq": {"expression": {"columnPath": "IntEsq"}},
                }
            },
            "filters": {
                "filterType": 6,
                "items": {
                    "IntNameFilter": {
                        "filterType": 1,
                        "comparisonType": 3,
                        "leftExpression": {"expressionType": 0, "columnPath": "IntName"},
                        "rightExpression": {
                            "expressionType": 2,
                            "parameter": {"dataValueType": 1, "value": int_name},
                        },
                    }
                },
            },
        }

    data = _select_query(session, headers, esq_body)
    rows = data.get("rows", [])
    if not rows:
        raise RuntimeError(f"IntExcelReport not found (code={report_code!r})")
    return rows[0]


def _resolve_year_month_id(session: requests.Session, headers: Dict[str, str], year_month_name: str) -> str:
    if not year_month_name or year_month_name.upper() == "__NONE__":
        return EMPTY_GUID

    esq_body = {
        "rootSchemaName": "BGYearMonth",
        "operationType": 0,
        "columns": {
            "items": {
                "Id": {"expression": {"columnPath": "Id"}},
                "Name": {"expression": {"columnPath": "Name"}},
            }
        },
        "filters": {
            "filterType": 6,
            "items": {
                "NameFilter": {
                    "filterType": 1,
                    "comparisonType": 3,
                    "leftExpression": {"expressionType": 0, "columnPath": "Name"},
                    "rightExpression": {
                        "expressionType": 2,
                        "parameter": {"dataValueType": 1, "value": year_month_name},
                    },
                }
            },
        },
    }
    data = _select_query(session, headers, esq_body)
    rows = data.get("rows", [])
    if not rows:
        print(f"WARN: YearMonth {year_month_name!r} not found; proceeding unfiltered")
        return EMPTY_GUID
    return rows[0].get("Id") or EMPTY_GUID


def _download_report(
    session: requests.Session,
    headers: Dict[str, str],
    key: str,
    candidates: List[str],
) -> Tuple[str, str, bytes, Optional[str]]:
    """Download the report bytes.

    Prefer the v8-first wrapper endpoint (UsrExcelReportService/GetReport) so the filename is .xlsm.
    Fall back to IntExcelReportService/GetReport for older deployments.

    Returns: (service_name, segment_used, bytes, content_disposition)
    """

    attempts: List[Tuple[str, str, int]] = []
    last_body: Optional[str] = None

    for seg in candidates:
        if not seg:
            continue

        for service_name in ("UsrExcelReportService", "IntExcelReportService"):
            url = f"{CREATIO_URL}/0/rest/{service_name}/GetReport/{key}/{seg}"
            resp = session.get(url, headers=headers, timeout=TIMEOUT_SECONDS)
            attempts.append((service_name, seg, resp.status_code))
            if resp.status_code == 200:
                content = resp.content
                if content[:2] == b"PK":
                    return service_name, seg, content, resp.headers.get("Content-Disposition")
                # Rare, but if we got 200 without a ZIP workbook, keep trying.
                last_body = resp.text[:300]
                continue
            last_body = resp.text[:300]

    summary = ", ".join([f"{svc}:{seg}:{status}" for svc, seg, status in attempts])
    raise RuntimeError(f"Download failed for all candidates. Attempts: {summary}. Last body: {last_body!r}")


def _count_non_empty_rows_xlsx(xlsx_path: Path, sheet_name: str) -> Optional[int]:
    if load_workbook is None:
        return None

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        non_empty = 0
        for row in ws.iter_rows(values_only=True):
            if any(v not in (None, "") for v in row):
                non_empty += 1
        return non_empty
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main() -> None:
    if not USERNAME or not PASSWORD:
        raise SystemExit("Set CREATIO_USERNAME and CREATIO_PASSWORD in your environment")

    session = requests.Session()

    # Login
    print("=== Logging in ===")
    resp = session.post(
        f"{CREATIO_URL}/ServiceModel/AuthService.svc/Login",
        json={"UserName": USERNAME, "UserPassword": PASSWORD},
        timeout=TIMEOUT_SECONDS,
    )
    print("Login:", "OK" if resp.status_code == 200 else f"FAILED ({resp.status_code})")

    headers = _headers(session)

    # Resolve report
    report = _resolve_int_excel_report(session, headers, REPORT_CODE)
    report_id = report.get("Id")
    int_name = report.get("IntName")
    entity_schema = report.get("IntEntitySchemaName")
    int_sheet = report.get("IntSheetName") or "Data"
    int_esq = report.get("IntEsq")

    print("\n=== Report ===")
    print(f"ReportCode: {REPORT_CODE}")
    print(f"ReportId: {report_id}")
    print(f"IntName: {int_name}")
    print(f"IntEntitySchemaName: {entity_schema}")
    print(f"IntSheetName: {int_sheet}")
    print(f"IntEsq length: {len(int_esq) if int_esq else 0}")

    # Resolve YearMonth
    year_month_id = _resolve_year_month_id(session, headers, YEAR_MONTH_NAME)
    if YEAR_MONTH_NAME and YEAR_MONTH_NAME.upper() != "__NONE__":
        print(f"\n=== YearMonth ===")
        print(f"Name: {YEAR_MONTH_NAME}")
        print(f"Id: {year_month_id}")
    else:
        print("\n=== YearMonth ===")
        print("(skipping filter)")

    sales_group_id = SALES_GROUP_ID or EMPTY_GUID

    # Generate
    print("\n=== Calling UsrExcelReportService.Generate ===")
    generate_url = f"{CREATIO_URL}/0/rest/UsrExcelReportService/Generate"
    generate_body = {
        "ReportId": report_id,
        "YearMonthId": year_month_id,
        # Note: frontend uses SalesRepId field to carry SalesGroupId (legacy naming).
        "SalesRepId": sales_group_id,
        "ExecutionId": EMPTY_GUID,
        "RecordCollection": [],
    }
    print(f"Request body: {json.dumps(generate_body, indent=2)}")

    resp = session.post(generate_url, json=generate_body, headers=headers, timeout=max(300, TIMEOUT_SECONDS))
    print(f"Response status: {resp.status_code}")
    if resp.status_code != 200:
        raise SystemExit(resp.text[:500])

    result = resp.json()
    print(f"Response: {json.dumps(result, indent=2)}")

    if not (result.get("success") and result.get("key")):
        raise SystemExit(f"Generate failed: {result.get('message')}")

    key = result["key"]

    # Download
    print("\n=== Downloading report ===")
    candidates: List[str] = []

    def _add(seg: str) -> None:
        seg = (seg or "").strip()
        if seg and seg not in candidates:
            candidates.append(seg)

    _add(REPORT_NAME_SEGMENT_OVERRIDE)
    _add(REPORT_CODE)
    _add(int_name or "")
    if isinstance(int_name, str) and int_name.startswith("Rpt "):
        _add(int_name[4:])

    svc, seg, content, cd = _download_report(session, headers, key, candidates)
    print(f"Downloaded via service={svc} reportNameSegment={seg!r}")
    if cd:
        print(f"Content-Disposition: {cd}")

    # Macro detection (DL-003): choose the correct extension.
    has_vba: Optional[bool] = None
    try:
        import zipfile
        from io import BytesIO

        with zipfile.ZipFile(BytesIO(content), "r") as z:
            has_vba = any(n.lower().endswith("xl/vbaproject.bin") for n in z.namelist())
    except Exception as e:
        print(f"WARN: macro detection failed: {e}")

    if has_vba is not None:
        print(f"Has macros (xl/vbaProject.bin): {has_vba}")

    ext = "xlsm" if has_vba else "xlsx"
    filename = ARTIFACTS_DIR / f"report_{REPORT_CODE}_{key[-8:]}.{ext}"
    filename.write_bytes(content)

    print(f"Saved to: {filename}")
    print(f"File size: {len(content)} bytes")
    print(f"Signature: {content[:2]!r} (expected b'PK' for xlsx/xlsm)")

    # Optional workbook inspection
    non_empty_rows = _count_non_empty_rows_xlsx(filename, int_sheet)
    if non_empty_rows is None:
        print("openpyxl not installed; skipping Excel inspection")
    else:
        print(f"Sheet {int_sheet!r}: non-empty rows = {non_empty_rows}")


if __name__ == "__main__":
    main()
